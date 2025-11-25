import os
import uuid
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Optional

from django.conf import settings
from django.db import transaction
from django.utils.text import slugify
from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple

from .models import MedicamentoCatalogo


def _valor_celda(ws, row: int, col: int):
    cell = ws.cell(row=row, column=col)
    return cell.value


def _anchor_row_col(image) -> (Optional[int], Optional[int]):
    anchor = getattr(image, "anchor", None)
    if isinstance(anchor, str):
        row, col = coordinate_to_tuple(anchor)
        return row, col
    anchor_obj = getattr(anchor, "_from", None) or getattr(anchor, "from", None)
    if anchor_obj and hasattr(anchor_obj, "row") and hasattr(anchor_obj, "col"):
        return anchor_obj.row + 1, anchor_obj.col + 1
    return None, None


def _guardar_imagen(img, nombre: str) -> str:
    data = img._data()
    filename = f"{slugify(nombre or 'medicamento')}-{uuid.uuid4().hex}.png"
    relative_path = os.path.join("medicamentos", filename)
    full_path = os.path.join(settings.MEDIA_ROOT, relative_path)
    os.makedirs(os.path.dirname(full_path), exist_ok=True)
    with open(full_path, "wb") as f:
        f.write(data)
    return relative_path


def parsear_catalogo_excel(ruta_excel: str) -> List[Dict]:
    wb = load_workbook(ruta_excel, data_only=True)
    ws = wb.active

    imagenes_por_fila = {}
    for img in getattr(ws, "_images", []):
        row, _ = _anchor_row_col(img)
        if row:
            imagenes_por_fila.setdefault(row, img)

    productos: List[Dict] = []
    fila = 1
    max_filas = ws.max_row or 0

    while fila <= max_filas:
        nombre = _valor_celda(ws, fila, 1)
        if not nombre or not str(nombre).strip():
            fila += 1
            continue

        codigo = _valor_celda(ws, fila + 1, 3)
        existencia = _valor_celda(ws, fila + 1, 7)
        departamento = _valor_celda(ws, fila + 2, 3)
        precio = _valor_celda(ws, fila + 2, 7)
        categoria = _valor_celda(ws, fila + 3, 3)

        imagen_rel = None
        for posible_fila in range(fila, fila + 5):
            if posible_fila in imagenes_por_fila:
                imagen_rel = _guardar_imagen(imagenes_por_fila[posible_fila], nombre)
                break

        productos.append(
            {
                "nombre": str(nombre).strip(),
                "codigo_barras": str(codigo).strip() if codigo else "",
                "existencia": existencia,
                "departamento": departamento or "",
                "categoria": categoria or "",
                "precio": precio,
                "imagen": imagen_rel,
            }
        )

        fila += 4

    return productos


def _limpiar_valor_decimal(valor) -> Decimal:
    try:
        if valor in [None, ""]:
            return Decimal("0")
        return Decimal(str(valor)).quantize(Decimal("0.01"))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal("0")


def _limpiar_valor_entero(valor) -> int:
    try:
        return int(valor)
    except (TypeError, ValueError):
        return 0


def actualizar_inventario(datos: List[Dict]) -> int:
    productos_creados = 0
    with transaction.atomic():
        MedicamentoCatalogo.objects.all().delete()
        nuevos = []
        codigos_vistos = set()

        for item in datos:
            nombre = (item.get("nombre") or "").strip() or "Producto sin nombre"
            codigo = (item.get("codigo_barras") or "").strip()
            if not codigo:
                codigo = f"codigo-{uuid.uuid4().hex[:8]}"
            if codigo in codigos_vistos:
                continue
            codigos_vistos.add(codigo)

            existencia = _limpiar_valor_entero(item.get("existencia"))
            precio = _limpiar_valor_decimal(item.get("precio"))
            departamento = (item.get("departamento") or "").strip()
            categoria = (item.get("categoria") or "").strip()
            imagen = item.get("imagen")

            med = MedicamentoCatalogo(
                nombre=nombre,
                codigo_barras=codigo,
                existencia=existencia,
                departamento=departamento,
                categoria=categoria,
                precio=precio,
            )

            if imagen:
                med.imagen = imagen

            nuevos.append(med)

        MedicamentoCatalogo.objects.bulk_create(nuevos)
        productos_creados = len(nuevos)

    return productos_creados
