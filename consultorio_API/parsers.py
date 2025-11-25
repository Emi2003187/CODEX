import uuid
from decimal import Decimal, InvalidOperation

from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from openpyxl import load_workbook


def _image_bytes(image_obj):
    data = image_obj._data if isinstance(getattr(image_obj, "_data", None), (bytes, bytearray)) else None
    if callable(getattr(image_obj, "_data", None)):
        data = image_obj._data()
    return data


def _guardar_imagen(image_obj):
    data = _image_bytes(image_obj)
    if not data:
        return None

    filename = f"medicamentos/{uuid.uuid4().hex}.png"
    default_storage.save(filename, ContentFile(data))
    return filename


def _obtener_mapa_imagenes(workbook):
    mapa = {}
    for imagen in getattr(workbook, "_images", []):
        anchor = getattr(imagen, "anchor", None)
        fila = None
        if hasattr(anchor, "_from"):
            fila = anchor._from.row + 1
        elif hasattr(anchor, "row"):
            fila = anchor.row
        if fila:
            mapa.setdefault(fila, []).append(imagen)
    return mapa


def _a_entero(valor):
    try:
        return int(valor)
    except (TypeError, ValueError):
        return 0


def _a_decimal(valor):
    try:
        return Decimal(str(valor))
    except (TypeError, ValueError, InvalidOperation):
        return None


def parsear_catalogo_excel(ruta):
    workbook = load_workbook(ruta, data_only=True)
    hoja = workbook.active

    imagenes_por_fila = _obtener_mapa_imagenes(workbook)
    productos = []

    fila = 1
    max_fila = hoja.max_row or 0

    while fila <= max_fila:
        nombre = hoja.cell(row=fila, column=1).value
        if nombre and isinstance(nombre, str) and nombre.strip():
            fila_clave = fila + 1
            fila_departamento = fila + 2
            fila_categoria = fila + 3

            codigo_barras = hoja.cell(row=fila_clave, column=3).value
            existencia = _a_entero(hoja.cell(row=fila_clave, column=7).value)

            departamento = hoja.cell(row=fila_departamento, column=3).value
            precio = _a_decimal(hoja.cell(row=fila_departamento, column=7).value)

            categoria = hoja.cell(row=fila_categoria, column=3).value

            imagen_path = None
            for candidate in [fila, fila_clave, fila_departamento, fila_categoria]:
                if candidate in imagenes_por_fila and imagenes_por_fila[candidate]:
                    imagen_path = _guardar_imagen(imagenes_por_fila[candidate][0])
                    break

            productos.append({
                "nombre": nombre.strip(),
                "codigo_barras": codigo_barras if codigo_barras else None,
                "existencia": existencia,
                "departamento": departamento if departamento else None,
                "categoria": categoria if categoria else None,
                "precio": precio,
                "imagen": imagen_path,
            })

            fila = fila_categoria + 1
        else:
            fila += 1

    return productos
