
import json
import os
import tempfile
import uuid
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.http import (
    FileResponse,
    HttpResponseForbidden,
    JsonResponse,
)
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.text import slugify
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST
from django.views.generic import DetailView
from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple

from .catalogo_excel import (
    buscar_articulos,
    catalogo_disponible,
    limpiar_cache_catalogo,
)
from .forms import ExcelUploadForm
from .models import MedicamentoCatalogo, MedicamentoRecetado, Receta


class _RecetaPDFBase(LoginRequiredMixin, DetailView):
    model = Receta
    pk_url_kwarg = "pk"

    def get(self, request, *args, **kwargs):
        receta = self.get_object()
        if receta.consulta.estado != "finalizada":
            messages.error(
                request,
                "La receta solo puede emitirse cuando la consulta está finalizada.",
            )
            return redirect("consulta_detalle", pk=receta.consulta.pk)

        buf = BytesIO()
        build_receta_pdf(buf, receta)
        buf.seek(0)
        fecha = receta.fecha_emision or timezone.now()
        filename = f"{receta.pk}_{slugify(receta.consulta.paciente.nombre_completo)}_{fecha.strftime('%Y%m%d')}.pdf"
        return FileResponse(buf, as_attachment=False, filename=filename, content_type="application/pdf")


class RecetaPreviewView(_RecetaPDFBase):
    """Previsualización/impresión de receta generada con ReportLab."""

    def dispatch(self, request, *args, **kwargs):
        receta = self.get_object()
        if not (
            request.user.has_perm("consultorio.view_receta")
            or request.user == receta.consulta.medico
        ):
            return HttpResponseForbidden()
        return super().dispatch(request, *args, **kwargs)


class RxRecetaView(RecetaPreviewView):
    pass


class RecetaA5View(_RecetaPDFBase):
    pass


def receta_pdf_reportlab(request, pk: int):
    receta = Receta.objects.select_related(
        "consulta", "consulta__paciente", "consulta__medico"
    ).prefetch_related("medicamentos").get(pk=pk)

    if not (
        request.user.has_perm("consultorio.view_receta")
        or request.user == getattr(receta.consulta, "medico", None)
    ):
        return HttpResponseForbidden()

    if receta.consulta.estado != "finalizada":
        messages.error(
            request,
            "La receta solo puede emitirse cuando la consulta está finalizada.",
        )
        return redirect("consulta_detalle", pk=receta.consulta.pk)

    buf = BytesIO()
    build_receta_pdf(buf, receta)
    buf.seek(0)
    fecha = receta.fecha_emision or timezone.now()
    filename = f"{receta.pk}_{slugify(receta.consulta.paciente.nombre_completo)}_{fecha.strftime('%Y%m%d')}.pdf"
    return FileResponse(buf, as_attachment=False, filename=filename, content_type="application/pdf")


@login_required
def cargar_excel_medicamentos(request):
    if request.user.rol != "admin":
        messages.error(request, "Solo un administrador puede acceder a este módulo.")
        return redirect("home")

    actualizados = None
    progress_points = []
    total_rows = 0

    if request.method == "POST":
        form = ExcelUploadForm(request.POST, request.FILES)

        if form.is_valid():
            archivo = request.FILES["archivo"]

            with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(archivo.name)[1]) as tmp:
                for chunk in archivo.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name

            datos = parsear_catalogo_excel(tmp_path)
            actualizados = actualizar_inventario(datos)
            total_rows = len(datos)

            if total_rows:
                progress_points = [int((i / total_rows) * 100) for i in range(1, total_rows + 1)]
                if progress_points[-1] < 100:
                    progress_points.append(100)

            messages.success(request, f"Actualizados: {actualizados}")

            try:
                os.remove(tmp_path)
            except OSError:
                pass
    else:
        form = ExcelUploadForm()

    return render(
        request,
        "PAGES/medicamentos/cargar_excel.html",
        {
            "form": form,
            "usuario": request.user,
            "actualizados": actualizados,
            "total_rows": total_rows,
            "progress_points": json.dumps(progress_points),
        }
    )


def _valor_celda(ws, row, col):
    cell = ws.cell(row=row, column=col)
    return cell.value


def _anchor_row_col(image):
    anchor = getattr(image, "anchor", None)
    if isinstance(anchor, str):
        row, col = coordinate_to_tuple(anchor)
        return row, col
    anchor_obj = getattr(anchor, "_from", None) or getattr(anchor, "from", None)
    if anchor_obj and hasattr(anchor_obj, "row") and hasattr(anchor_obj, "col"):
        return anchor_obj.row + 1, anchor_obj.col + 1
    return None, None


def _guardar_imagen(img, nombre):
    data = img._data()
    filename = f"{slugify(nombre or 'medicamento')}-{uuid.uuid4().hex}.png"
    relative_path = os.path.join("medicamentos", filename)
    full_path = os.path.join(settings.MEDIA_ROOT, relative_path)
    os.makedirs(os.path.dirname(full_path), exist_ok=True)
    with open(full_path, "wb") as f:
        f.write(data)
    return relative_path


def parsear_catalogo_excel(ruta_excel):
    wb = load_workbook(ruta_excel, data_only=True)
    ws = wb.active

    imagenes_por_fila = {}
    for img in getattr(ws, "_images", []):
        row, _ = _anchor_row_col(img)
        if row:
            imagenes_por_fila.setdefault(row, img)

    productos = []
    fila = 1
    max_filas = ws.max_row or 0

    while fila <= max_filas:
        nombre = _valor_celda(ws, fila, 1)
        if not nombre or not str(nombre).strip():
            fila += 1
            continue

        codigo = _valor_celda(ws, fila + 1, 3)
        existencia = _valor_celda(ws, fila + 1, 7)
        departamento = _valor_celda(ws, fila + 2, 3)
        precio = _valor_celda(ws, fila + 2, 7)
        categoria = _valor_celda(ws, fila + 3, 3)

        imagen_rel = None
        for posible_fila in range(fila, fila + 5):
            if posible_fila in imagenes_por_fila:
                imagen_rel = _guardar_imagen(imagenes_por_fila[posible_fila], nombre)
                break

        productos.append(
            {
                "nombre": str(nombre).strip(),
                "codigo_barras": str(codigo).strip() if codigo else "",
                "existencia": existencia,
                "departamento": departamento or "",
                "categoria": categoria or "",
                "precio": precio,
                "imagen": imagen_rel,
            }
        )

        fila += 4

    return productos


def actualizar_inventario(datos):
    productos_creados = 0
    with transaction.atomic():
        MedicamentoCatalogo.objects.all().delete()
        nuevos = []

        for item in datos:
            nombre = (item.get("nombre") or "").strip() or "Producto sin nombre"
            codigo = (item.get("codigo_barras") or "").strip() or f"codigo-{uuid.uuid4().hex[:8]}"
            try:
                existencia = int(item.get("existencia") or 0)
            except (TypeError, ValueError):
                existencia = 0

            try:
                precio_valor = item.get("precio")
                precio = Decimal(str(precio_valor)) if precio_valor not in [None, ""] else Decimal("0")
            except (InvalidOperation, TypeError, ValueError):
                precio = Decimal("0")

            departamento = (item.get("departamento") or "").strip()
            categoria = (item.get("categoria") or "").strip()
            imagen = item.get("imagen")

            med = MedicamentoCatalogo(
                nombre=nombre,
                codigo_barras=codigo,
                existencia=existencia,
                departamento=departamento,
                categoria=categoria,
                precio=precio,
            )

            if imagen:
                med.imagen = imagen

            nuevos.append(med)

        MedicamentoCatalogo.objects.bulk_create(nuevos)
        productos_creados = len(nuevos)

    return productos_creados


@login_required
@require_GET
def catalogo_excel_json(request):
    q = (request.GET.get("q") or "").strip()
    page = int(request.GET.get("page") or 1)
    per_page = int(request.GET.get("per_page") or 15)

    data = buscar_articulos(q=q, page=page, per_page=per_page)
    for it in data.get("items", []):
        if it.get("imagen_url"):
            it["imagen"] = it["imagen_url"]
        it.pop("imagen_url", None)

    return JsonResponse(data)


@csrf_exempt
@login_required
@require_POST
def catalogo_excel_limpiar_cache(request):
    """Limpia la caché del catálogo de medicamentos."""
    limpiar_cache_catalogo()
    return JsonResponse({"ok": True})


@login_required
@require_GET
def receta_catalogo_excel(request, receta_id):
    receta = get_object_or_404(Receta, id=receta_id)
    return render(
        request,
        "PAGES/recetas/catalogo_excel.html",
        {"receta": receta, "excel_disponible": catalogo_disponible()},
    )


@login_required
@require_POST
@transaction.atomic
def receta_catalogo_excel_agregar(request, receta_id):
    receta = get_object_or_404(Receta, id=receta_id)
    nombre = (request.POST.get("nombre") or "").strip()
    clave = (request.POST.get("clave") or "").strip()
    dosis = (request.POST.get("dosis") or "").strip()
    frecuencia = (request.POST.get("frecuencia") or "").strip()
    via_administracion = (request.POST.get("via_administracion") or "").strip()
    indicaciones_especificas = (request.POST.get("indicaciones_especificas") or "").strip()
    try:
        cantidad = int(request.POST.get("cantidad") or "1")
    except Exception:
        cantidad = 1
    if cantidad < 1:
        cantidad = 1
    if not nombre and not clave:
        return JsonResponse({"ok": False, "error": "Nombre requerido"}, status=400)
    cat = None
    if clave or nombre:
        data = buscar_articulos(q=clave or nombre, page=1, per_page=1)
        items = data.get("items", [])
        if items:
            cat = items[0]

    cat_nombre = cat.get("nombre") if cat else nombre

    mr = None
    if clave:
        mr = MedicamentoRecetado.objects.filter(
            receta=receta, codigo_barras=clave
        ).first()
    if not mr:
        mr = MedicamentoRecetado.objects.filter(
            receta=receta, nombre=cat_nombre
        ).first()

    if mr:
        mr.cantidad += cantidad
        if dosis:
            mr.dosis = dosis
        if frecuencia:
            mr.frecuencia = frecuencia
        if via_administracion:
            mr.via_administracion = via_administracion
        if indicaciones_especificas:
            mr.indicaciones_especificas = indicaciones_especificas
        if not mr.codigo_barras and clave:
            mr.codigo_barras = clave
        if cat:
            mr.existencia = cat.get("existencia", mr.existencia)
            mr.categoria = cat.get("categoria", mr.categoria)
            mr.departamento = cat.get("departamento", mr.departamento)
        mr.save()
    else:
        mr = MedicamentoRecetado.objects.create(
            receta=receta,
            nombre=cat_nombre,
            cantidad=cantidad,
            dosis=dosis,
            frecuencia=frecuencia,
            via_administracion=via_administracion or None,
            indicaciones_especificas=indicaciones_especificas or None,
            existencia=cat.get("existencia", 0) if cat else 0,
            codigo_barras=cat.get("clave") if cat else (clave or None),
            categoria=cat.get("categoria") if cat else None,
            departamento=cat.get("departamento") if cat else None,
        )

    return JsonResponse(
        {
            "ok": True,
            "id": str(mr.id),
            "nombre": mr.nombre,
            "cantidad": mr.cantidad,
            "codigo_barras": mr.codigo_barras or "",
            "principio_activo": mr.principio_activo or "",
            "dosis": mr.dosis,
            "frecuencia": mr.frecuencia,
            "via_administracion": mr.via_administracion or "",
            "indicaciones_especificas": mr.indicaciones_especificas or "",
        }
    )


@login_required
@require_GET
def receta_medicamentos_json(request, receta_id):
    """Devuelve los medicamentos actuales de la receta."""
    receta = get_object_or_404(Receta, id=receta_id)
    items = [
        {
            "id": mr.id,
            "nombre": mr.nombre,
            "principio_activo": mr.principio_activo or "",
            "cantidad": mr.cantidad,
            "codigo_barras": mr.codigo_barras or "",
            "dosis": mr.dosis,
            "frecuencia": mr.frecuencia,
            "via_administracion": mr.via_administracion or "",
            "indicaciones_especificas": mr.indicaciones_especificas or "",
        }
        for mr in receta.medicamentos.all()
    ]
    return JsonResponse({"items": items})


@login_required
@require_POST
@transaction.atomic
def receta_medicamento_actualizar(request, receta_id, med_id):
    """Actualiza los campos de un medicamento en la receta."""
    receta = get_object_or_404(Receta, id=receta_id)
    med = get_object_or_404(MedicamentoRecetado, id=med_id, receta=receta)

    try:
        cantidad = int(request.POST.get("cantidad") or "1")
    except Exception:
        cantidad = 1
    if cantidad < 1:
        cantidad = 1
    med.cantidad = cantidad

    dosis = request.POST.get("dosis")
    if dosis is not None:
        med.dosis = dosis

    frecuencia = request.POST.get("frecuencia")
    if frecuencia is not None:
        med.frecuencia = frecuencia

    via_administracion = request.POST.get("via_administracion")
    if via_administracion is not None:
        med.via_administracion = via_administracion or None

    indicaciones = request.POST.get("indicaciones_especificas")
    if indicaciones is not None:
        med.indicaciones_especificas = indicaciones or None

    med.save()
    return JsonResponse(
        {
            "ok": True,
            "id": med.id,
            "cantidad": med.cantidad,
            "dosis": med.dosis,
            "frecuencia": med.frecuencia,
            "via_administracion": med.via_administracion or "",
            "indicaciones_especificas": med.indicaciones_especificas or "",
        }
    )


@login_required
@require_POST
@transaction.atomic
def receta_medicamento_eliminar(request, receta_id, med_id):
    """Elimina un medicamento de la receta."""
    receta = get_object_or_404(Receta, id=receta_id)
    med = get_object_or_404(MedicamentoRecetado, id=med_id, receta=receta)
    med.delete()
    return JsonResponse({"ok": True})
