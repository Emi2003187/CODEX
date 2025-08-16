# consultorio_API/catalogo_excel.py
# -*- coding: utf-8 -*-
"""
Extractor robusto del catálogo Excel para Recetas.

Formato esperado (por bloques, con fila en blanco entre artículos):

  <Nombre / Presentación>         ← línea sin “:”
  Clave: <...>      Existencia: <...>
  Departamento: <...>   Precio: <...>
  Categoría: <...>
  (puede haber una imagen embebida cerca)

Funciones clave:
- Lee TODAS las hojas.
- Detecta cada artículo anclándose en la etiqueta “Clave:”.
- Nombre: sube hasta 5 filas y toma la última línea sin “:”, ignorando encabezados.
- Resto de etiquetas: busca valor en ventana (misma fila y hasta 10 filas abajo; +60 columnas).
- Extrae imágenes embebidas desde xl/media y las mapea por fila usando drawings.
"""

from __future__ import annotations

import os
import hashlib
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional
from unicodedata import normalize
from zipfile import ZipFile
from xml.etree import ElementTree as ET

from django.conf import settings

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None


# ───────────────────────── Ruta al Excel ─────────────────────────
def _find_excel_path() -> Path:
    base = Path(getattr(settings, "BASE_DIR", "."))
    cfg = getattr(settings, "CATALOGO_EXCEL_PATH", None)
    if cfg:
        p = Path(cfg)
        if p.exists():
            return p
    candidates = [
        "Catalogo de Artículos.xlsx",
        "Catalogo de Articulos.xlsx",
        "Catálogo de Artículos.xlsx",
        "Catalogo de Articulos.xlsm",
        "Catalogo de Artículos.xlsm",
    ]
    for name in candidates:
        p = base / name
        if p.exists():
            return p
    for p in base.glob("Catalogo de Art*.*xls*"):
        return p
    return base / "Catalogo de Artículos.xlsx"


EXCEL_PATH = _find_excel_path()


# ─────────────────────── Utilidades de texto/número ────────────────────────
def _strip_nbsp(s: str) -> str:
    return str(s).replace("\xa0", " ").replace("\u202f", " ").strip()


def _norm_text(s: str) -> str:
    if s is None:
        return ""
    s = _strip_nbsp(s)
    s = normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return s.lower().strip()


def _toi(v) -> int:
    if v is None:
        return 0
    s = str(v).strip().replace(",", "")
    try:
        return int(s)
    except Exception:
        try:
            return int(float(s))
        except Exception:
            return 0


def _tof(v) -> float:
    if v is None:
        return 0.0
    s = str(v).strip().replace("$", "").replace(",", "")
    try:
        return float(s)
    except Exception:
        return 0.0


# ────────────────────────── Búsquedas en la hoja ────────────────────────────
def _first_nonempty_to_right(
    grid: List[List[str]], r: int, c: int, max_down: int = 10, max_right: int = 60
) -> str:
    """
    Primer valor NO vacío a la derecha del (r,c) en filas r..r+max_down.
    Aumentado: mira hasta 10 filas abajo y 60 columnas a la derecha.
    """
    rows = len(grid)
    for rr in range(r, min(rows, r + 1 + max_down)):
        row = grid[rr]
        for cc in range(c + 1, min(len(row), c + 1 + max_right)):
            raw = _strip_nbsp(row[cc])
            if raw:
                return raw
    return ""


def _first_numeric_to_right(
    grid: List[List[str]], r: int, c: int, max_down: int = 10, max_right: int = 60
) -> str:
    """
    Primer valor con dígitos a la derecha del (r,c) en filas r..r+max_down.
    Aumentado: mira hasta 10 filas abajo y 60 columnas a la derecha.
    """
    rows = len(grid)
    for rr in range(r, min(rows, r + 1 + max_down)):
        row = grid[rr]
        for cc in range(c + 1, min(len(row), c + 1 + max_right)):
            raw = _strip_nbsp(row[cc])
            if raw and any(ch.isdigit() for ch in raw):
                return raw
    return ""


def _find_label_value(
    grid: List[List[str]], r_from: int, label: str, numeric: bool
) -> str:
    """
    Busca una celda cuyo texto normalizado empiece con `label`
    en la fila r_from o hasta 10 filas por debajo; devuelve el valor
    a la derecha (no vacío o numérico según `numeric`) mirando hasta
    60 columnas a la derecha.
    """
    rows = len(grid)
    MAX_DOWN = 10
    MAX_RIGHT = 60

    for rr in range(r_from, min(rows, r_from + 1 + MAX_DOWN)):
        row = grid[rr]
        for cc, val in enumerate(row):
            if not val:
                continue
            tag = _norm_text(val)
            if tag.startswith(label):  # p.ej. "categoria", "departamento", "precio"
                return (
                    _first_numeric_to_right if numeric else _first_nonempty_to_right
                )(grid, rr, cc, max_down=MAX_DOWN, max_right=MAX_RIGHT)
    return ""


def _find_name_above(grid: List[List[str]], r_from: int) -> str:
    """
    Busca hacia ARRIBA (hasta 5 filas) la última línea no vacía SIN “:”
    y que no sea un encabezado como “catalogo de articulos” o “farmacia nova”.
    """
    ignore = {"catalogo de articulos", "farmacia nova"}
    for rr in range(max(0, r_from - 5), r_from)[::-1]:
        row = grid[rr]
        joined = " ".join([c for c in row if c]).strip()
        if joined and ":" not in joined:
            if _norm_text(joined) not in ignore:
                return joined
    return ""


def _append(
    items: List[Dict[str, Any]],
    nombre,
    clave,
    existencia,
    dep,
    precio,
    cat,
    img=None,
) -> None:
    raw_clave = str(clave or "").strip()
    if not raw_clave:
        return
    clave_clean = "".join(ch for ch in raw_clave if ch.isdigit()) or raw_clave
    items.append(
        {
            "nombre": str(nombre or "").strip(),
            "clave": clave_clean,
            "existencia": _toi(existencia),
            "departamento": str(dep or "").strip(),
            "precio": _tof(precio),
            "categoria": str(cat or "").strip(),
            "imagen_url": str(img or "").strip(),
        }
    )


# ──────────────────────── Extracción de imágenes ────────────────────────────
_NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "ws": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
}


def _read_xml(zf: ZipFile, name: str) -> Optional[ET.Element]:
    try:
        return ET.fromstring(zf.read(name))
    except Exception:
        return None


def _extract_images_index(xlsx_path: Path) -> Dict[str, List[Tuple[int, str]]]:
    """
    Devuelve un índice: {nombre_hoja: [(row0, url_media), ...]}
    row0 es 0-based (fila ancla de la imagen en la hoja).
    """
    media_urls: Dict[str, List[Tuple[int, str]]] = {}
    if not xlsx_path.exists():
        return media_urls

    media_out_dir = Path(getattr(settings, "MEDIA_ROOT", ".")) / "catalogo_excel"
    media_out_dir.mkdir(parents=True, exist_ok=True)
    media_url_base = (
        getattr(settings, "MEDIA_URL", "/media/").rstrip("/") + "/catalogo_excel"
    )

    with ZipFile(xlsx_path) as zf:
        # 1) workbook: mapear sheets -> targets (sheetX.xml)
        wb = _read_xml(zf, "xl/workbook.xml")
        wb_rels = _read_xml(zf, "xl/_rels/workbook.xml.rels")
        if wb is None or wb_rels is None:
            return media_urls

        rels = {}
        for rel in wb_rels.findall(
            ".//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"
        ):
            rels[rel.attrib.get("Id")] = rel.attrib.get("Target")

        sheet_to_xml = {}  # sheet name -> 'xl/worksheets/sheetN.xml'
        for sh in wb.findall(".//ws:sheet", _NS):
            name = sh.attrib.get("name")
            rid = sh.attrib.get("{%s}id" % _NS["r"])
            target = rels.get(rid, "")
            if not target:
                continue
            if not target.startswith("xl/"):
                target = "xl/" + target.lstrip("./")
            sheet_to_xml[name] = target

        # 2) por cada sheet, localizar drawing asociado
        for sheet_name, sheet_xml in sheet_to_xml.items():
            rels_name = sheet_xml.replace(
                "xl/worksheets/", "xl/worksheets/_rels/"
            ) + ".rels"
            srels = _read_xml(zf, rels_name)
            if srels is None:
                continue

            drawing_target = None
            for rel in srels.findall(
                ".//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"
            ):
                if rel.attrib.get("Type", "").endswith("/drawing"):
                    drawing_target = rel.attrib.get("Target")
                    break
            if not drawing_target:
                continue
            if drawing_target.startswith("../"):
                drawing_xml = "xl/" + drawing_target.replace("../", "")
            elif not drawing_target.startswith("xl/"):
                drawing_xml = "xl/" + drawing_target.lstrip("./")
            else:
                drawing_xml = drawing_target

            # 3) relaciones del drawing → rId -> media/imageN.png
            d_rels_name = drawing_xml.replace(
                "xl/drawings/", "xl/drawings/_rels/"
            ) + ".rels"
            d_rels = _read_xml(zf, d_rels_name)
            if d_rels is None:
                continue
            dmap = {}
            for rel in d_rels.findall(
                ".//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"
            ):
                dmap[rel.attrib.get("Id")] = rel.attrib.get("Target")

            # 4) drawing xml: anclas y rIds
            dxml = _read_xml(zf, drawing_xml)
            if dxml is None:
                continue

            pairs: List[Tuple[int, str]] = []

            def _collect(two_or_one):
                _from = two_or_one.find("xdr:from", _NS)
                if _from is None:
                    return
                row_el = _from.find("xdr:row", _NS)
                if row_el is None or not row_el.text:
                    return
                row0 = int(row_el.text)  # 0-based

                blip = two_or_one.find(".//a:blip", _NS)
                if blip is None:
                    return
                rid = blip.attrib.get("{%s}embed" % _NS["r"])
                target = dmap.get(rid, "")
                if not target:
                    return
                # normaliza ruta del media
                if target.startswith("../"):
                    media_part = "xl/" + target.replace("../", "")
                elif target.startswith("media/"):
                    media_part = "xl/" + target
                else:
                    media_part = (
                        target if target.startswith("xl/") else "xl/" + target.lstrip("./")
                    )

                # lee bytes
                try:
                    data = zf.read(media_part)
                except Exception:
                    return

                # guarda en MEDIA_ROOT/catalogo_excel
                ext = os.path.splitext(media_part)[1].lower() or ".png"
                h = hashlib.sha1(
                    (sheet_name + media_part).encode("utf-8")
                ).hexdigest()[:16]
                fname = f"{sheet_name}_{row0}_{h}{ext}".replace(" ", "_")
                out_path = media_out_dir / fname
                if not out_path.exists():
                    with open(out_path, "wb") as f:
                        f.write(data)
                url = f"{media_url_base}/{fname}"
                pairs.append((row0, url))

            for node in dxml.findall("xdr:twoCellAnchor", _NS):
                _collect(node)
            for node in dxml.findall("xdr:oneCellAnchor", _NS):
                _collect(node)

            if pairs:
                media_urls[sheet_name] = pairs

    return media_urls


def _closest_image_for_row(
    sheet_images: List[Tuple[int, str]], r: int, max_delta: int = 5
) -> str:
    """
    Retorna la URL de la imagen cuya fila (0-based) es más cercana a r.
    Solo si la distancia es <= max_delta. Si no, devuelve "".
    """
    if not sheet_images:
        return ""
    best = None
    best_d = 10**9
    for row0, url in sheet_images:
        d = abs(row0 - r)
        if d < best_d:
            best_d, best = d, url
    return best if best is not None and best_d <= max_delta else ""


# ─────────────────────────── Parseo por hoja ────────────────────────────────
def _parse_sheet(ws, sheet_img_index: List[Tuple[int, str]]) -> List[Dict[str, Any]]:
    # hoja → grid de strings
    grid: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        grid.append([_strip_nbsp("" if v is None else v) for v in row])

    items: List[Dict[str, Any]] = []

    rows = len(grid)
    for r in range(rows):
        row = grid[r]
        for c, val in enumerate(row):
            if not val:
                continue
            tag = _norm_text(val)
            if tag.startswith("clave"):
                # NUEVO artículo
                nombre = _find_name_above(grid, r)
                clave = _first_nonempty_to_right(grid, r, c)

                existencia = _find_label_value(grid, r, "existencia", numeric=True)
                dep = _find_label_value(grid, r, "departamento", numeric=False)
                precio = _find_label_value(grid, r, "precio", numeric=True)
                categoria = _find_label_value(grid, r, "categoria", numeric=False)

                # Imagen más cercana por fila (±5 filas)
                imagen = _closest_image_for_row(sheet_img_index, r, max_delta=5)

                _append(
                    items, nombre, clave, existencia, dep, precio, categoria, imagen
                )

    return items


# ─────────────────────────── API pública ────────────────────────────────────
def catalogo_disponible() -> bool:
    return EXCEL_PATH.exists() and load_workbook is not None


def buscar_articulos(q: str = "", page: int = 1, per_page: int = 15) -> Dict[str, Any]:
    if not catalogo_disponible():
        return {"items": [], "total": 0, "page": 1, "per_page": per_page}

    # Índice de imágenes por (hoja → [(row0, url), ...])
    img_index_by_sheet = _extract_images_index(EXCEL_PATH)

    wb = load_workbook(EXCEL_PATH, data_only=True)

    all_items: List[Dict[str, Any]] = []
    for ws in wb.worksheets:
        sheet_imgs = img_index_by_sheet.get(ws.title, [])
        try:
            parsed = _parse_sheet(ws, sheet_imgs)
            if parsed:
                all_items.extend(parsed)
        except Exception:
            continue

    if not all_items:
        return {"items": [], "total": 0, "page": 1, "per_page": per_page}

    # Filtro
    if q:
        s = _norm_text(q)

        def ok(it: Dict[str, Any]) -> bool:
            return (
                s in _norm_text(it.get("nombre", ""))
                or s in _norm_text(it.get("clave", ""))
                or s in _norm_text(it.get("departamento", ""))
                or s in _norm_text(it.get("categoria", ""))
                or (
                    s.replace(".", "").isdigit()
                    and abs(it.get("precio", 0.0) - _tof(q)) < 1e-9
                )
            )

        all_items = [it for it in all_items if ok(it)]

    total = len(all_items)
    page = max(1, int(page or 1))
    per_page = int(per_page or 15)
    start = (page - 1) * per_page
    end = start + per_page

    return {"items": all_items[start:end], "total": total, "page": page, "per_page": per_page}


__all__ = [
    "EXCEL_PATH",
    "catalogo_disponible",
    "buscar_articulos",
]
