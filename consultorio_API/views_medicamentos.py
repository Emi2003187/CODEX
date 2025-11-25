from __future__ import annotations

import csv
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path
import re
from typing import Dict, List, Tuple

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.views import View
from openpyxl import load_workbook

from .forms import ExcelUploadForm, MedicamentoCatalogoForm
from .models import MedicamentoCatalogo


@dataclass
class MedicamentoParsed:
    nombre: str
    clave: str
    departamento: str | None
    categoria: str | None
    existencia: int
    precio: Decimal | None
    imagen_data: bytes | None = None
    imagen_ext: str | None = None


def _clean_line(row: List[object]) -> str:
    parts = []
    for cell in row:
        if cell is None:
            continue
        text = str(cell).strip()
        if text:
            parts.append(text)
    return " ".join(parts).strip()


def _strip_label_value(text: str, key: str | None = None) -> str:
    """Remove the label prefix from a line and return the raw value.

    The Excel export sometimes introduces extra spaces or omits the colon after
    the label, e.g. ``"Precio   $ 15.00"``. This helper trims the label keyword
    (if provided) and any leading punctuation/whitespace so value parsing
    remains robust.
    """

    cleaned = text.strip()
    if key:
        variants = [key]
        if key == "categoria":
            variants.append("categoría")

        expanded_variants = []
        for v in variants:
            expanded_variants.extend([v, f"{v}:", f"{v} :", f"{v} : "])

        lower_cleaned = cleaned.lower()
        for variant in expanded_variants:
            if lower_cleaned.startswith(variant):
                cleaned = cleaned[len(variant) :].strip()
                break

    if ":" in cleaned:
        cleaned = cleaned.split(":", 1)[1].strip()
    return cleaned


LABEL_REGEX = re.compile(
    r"(clave|departamento|categor[ií]a|existencia|precio)\s*:??\s*(.*?)\s*(?=(?:clave|departamento|categor[ií]a|existencia|precio)\s*:|$)",
    re.IGNORECASE,
)


def _extract_labels(text: str) -> Dict[str, str]:
    """Extrae múltiples pares etiqueta-valor de una misma línea.

    Soporta combinaciones como ``"Clave: 7501 Existencia: 4 Precio: $ 15.00"``
    sin perder ninguno de los campos.
    """

    found: Dict[str, str] = {}
    for match in LABEL_REGEX.finditer(text):
        key = match.group(1).lower()
        value = match.group(2).strip()
        if key.startswith("categor"):
            key = "categoria"
        found[key] = value
    return found


def _parse_int(value: str | None) -> int | None:
    if value is None:
        return None
    cleaned = value.replace(",", "").replace(" ", "").strip()
    lowered = cleaned.lower()
    if lowered in {"n/a", "na", "n.d", "nd", "na."}:
        return 0
    if not cleaned:
        return None
    try:
        return int(float(cleaned))
    except Exception:
        return None


def _parse_decimal(value: str | None) -> Decimal | None:
    if value is None:
        return None
    cleaned = value.replace("$", "").replace(",", "").replace(" ", "").strip()
    lowered = cleaned.lower()
    if lowered in {"n/a", "na", "n.d", "nd", "na."}:
        return Decimal(0)
    if not cleaned:
        return None
    try:
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None


def _extract_sheet_images(ws) -> Dict[int, List[Tuple[bytes, str | None]]]:
    """Devuelve un mapeo row_index -> lista de (bytes, extensión)."""

    images_by_row: Dict[int, List[Tuple[bytes, str | None]]] = {}

    for img in getattr(ws, "_images", []):  # pragma: no cover - dependiente de archivo
        try:
            anchor = getattr(img, "anchor", None)
            if hasattr(anchor, "_from"):
                row_idx = anchor._from.row
            elif hasattr(anchor, "row"):
                row_idx = anchor.row
            else:
                continue

            data = img._data() if hasattr(img, "_data") else None
            if not data:
                continue

            ext = getattr(img, "format", None)
            images_by_row.setdefault(row_idx, []).append((data, ext))
        except Exception:
            continue

    return images_by_row


def _pop_image_for_block(images_by_row: Dict[int, List[Tuple[bytes, str | None]]], start: int, end: int):
    if not images_by_row:
        return None, None

    for row_idx in range(start, end + 1):
        if row_idx in images_by_row and images_by_row[row_idx]:
            data, ext = images_by_row[row_idx].pop(0)
            if not images_by_row[row_idx]:
                images_by_row.pop(row_idx, None)
            return data, ext
    return None, None


def _parse_sheet(rows: List[List[object]], images_by_row: Dict[int, List[Tuple[bytes, str | None]]] | None = None) -> Tuple[List[MedicamentoParsed], List[dict]]:
    """Parsea una hoja en el formato visual del catálogo."""

    items: List[MedicamentoParsed] = []
    errors: List[dict] = []
    i = 0

    skip_prefixes = {"catalogo de articulos", "catálogo de artículos", "farmacia nova"}

    while i < len(rows):
        line = _clean_line(rows[i])
        if not line:
            i += 1
            continue

        if _extract_labels(line):
            i += 1
            continue

        lowered = line.lower()
        if any(lowered.startswith(prefix) for prefix in skip_prefixes):
            i += 1
            continue

        nombre = line
        info = {
            "clave": None,
            "departamento": None,
            "categoria": None,
            "existencia": None,
            "precio": None,
        }

        found_detail = False
        j = i + 1
        max_seek = min(len(rows), i + 25)
        next_candidate_index = None

        while j < max_seek:
            detalle_text = _clean_line(rows[j])
            if not detalle_text:
                j += 1
                continue

            labels_found = _extract_labels(detalle_text)
            if labels_found:
                found_detail = True
                for key, val in labels_found.items():
                    info[key] = _strip_label_value(val, key)
                j += 1
                continue

            next_candidate_index = j
            break

        missing_fields = [field for field in ["clave", "existencia", "precio"] if not info.get(field)]

        if missing_fields:
            if found_detail:
                errors.append(
                    {
                        "linea": i + 1,
                        "nombre": nombre,
                        "motivo": f"Faltan campos: {', '.join(missing_fields)}",
                    }
                )
            i = next_candidate_index if next_candidate_index is not None else j
            continue

        existencia_val = _parse_int(str(info["existencia"]))
        precio_val = _parse_decimal(str(info["precio"]))

        if existencia_val is None or precio_val is None:
            motivo = "Existencia inválida" if existencia_val is None else "Precio inválido"
            errors.append({"linea": i + 1, "nombre": nombre, "motivo": motivo})
            i = next_candidate_index if next_candidate_index is not None else j
            continue

        block_end = (next_candidate_index - 1) if next_candidate_index is not None else (j - 1)
        img_data, img_ext = _pop_image_for_block(images_by_row or {}, i, block_end)

        items.append(
            MedicamentoParsed(
                nombre=nombre,
                clave=str(info["clave"]).strip(),
                departamento=str(info["departamento"] or "").strip() or None,
                categoria=str(info["categoria"] or "").strip() or None,
                existencia=existencia_val,
                precio=precio_val,
                imagen_data=img_data,
                imagen_ext=img_ext,
            )
        )

        i = next_candidate_index if next_candidate_index is not None else j

    return items, errors


def parse_medications_file(file_bytes: bytes, filename: str) -> Tuple[List[MedicamentoParsed], List[dict]]:
    ext = Path(filename).suffix.lower()
    if ext == ".csv":
        text = file_bytes.decode("utf-8", errors="ignore")
        reader_rows = list(csv.reader(StringIO(text)))
        return _parse_sheet(reader_rows)

    if ext in {".xlsx", ".xls"}:
        try:
            wb = load_workbook(BytesIO(file_bytes), data_only=True)
        except Exception as exc:  # pragma: no cover - dependiente de archivo
            raise ValueError("No se pudo leer el archivo. Verifique el formato.") from exc

        items: List[MedicamentoParsed] = []
        errors: List[dict] = []
        for ws in wb.worksheets:
            sheet_rows = [list(row) for row in ws.iter_rows(values_only=True)]
            images_by_row = _extract_sheet_images(ws)
            sheet_items, sheet_errors = _parse_sheet(sheet_rows, images_by_row)
            items.extend(sheet_items)
            for err in sheet_errors:
                err["hoja"] = ws.title
                errors.append(err)
        return items, errors

    raise ValueError("Formato de archivo no soportado")


@login_required
def medicamentos_lista(request):
    termino = request.GET.get("q", "").strip()
    qs = MedicamentoCatalogo.objects.all()
    if termino:
        qs = qs.filter(
            Q(nombre__icontains=termino)
            | Q(clave__icontains=termino)
            | Q(categoria__icontains=termino)
            | Q(departamento__icontains=termino)
        )
    qs = qs.order_by("nombre")
    paginator = Paginator(qs, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "medicamentos/lista.html",
        {"page_obj": page_obj, "termino": termino, "usuario": request.user},
    )


@login_required
def medicamentos_crear(request):
    if request.method == "POST":
        form = MedicamentoCatalogoForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Medicamento creado correctamente.")
            return redirect("medicamentos_lista")
    else:
        form = MedicamentoCatalogoForm()

    return render(
        request,
        "medicamentos/formulario.html",
        {"form": form, "accion": "Crear", "usuario": request.user},
    )


@login_required
def medicamentos_editar(request, pk: int):
    medicamento = get_object_or_404(MedicamentoCatalogo, pk=pk)
    if request.method == "POST":
        form = MedicamentoCatalogoForm(request.POST, request.FILES, instance=medicamento)
        if form.is_valid():
            form.save()
            messages.success(request, "Medicamento actualizado correctamente.")
            return redirect("medicamentos_lista")
    else:
        form = MedicamentoCatalogoForm(instance=medicamento)

    return render(
        request,
        "medicamentos/formulario.html",
        {
            "form": form,
            "accion": "Editar",
            "medicamento": medicamento,
            "usuario": request.user,
        },
    )


@login_required
def medicamentos_eliminar(request, pk: int):
    medicamento = get_object_or_404(MedicamentoCatalogo, pk=pk)
    if request.method == "POST":
        medicamento.delete()
        messages.success(request, "Medicamento eliminado correctamente.")
        return redirect("medicamentos_lista")
    return render(
        request,
        "medicamentos/confirmar_eliminar.html",
        {"medicamento": medicamento, "usuario": request.user},
    )


class MedicamentoExcelUploadView(LoginRequiredMixin, View):
    template_name = "medicamentos/cargar_excel.html"

    def get(self, request):
        return render(
            request,
            self.template_name,
            {
                "form": ExcelUploadForm(),
                "creados": None,
                "actualizados": None,
                "errores": None,
                "total": None,
                "usuario": request.user,
            },
        )

    def post(self, request):
        form = ExcelUploadForm(request.POST, request.FILES)
        context = {
            "form": form,
            "creados": None,
            "actualizados": None,
            "errores": None,
            "total": None,
            "usuario": request.user,
        }
        if form.is_valid():
            archivo = form.cleaned_data["archivo"]
            file_bytes = archivo.read()
            try:
                items, errores = parse_medications_file(file_bytes, archivo.name)
            except ValueError as exc:
                messages.error(request, str(exc))
                return render(request, self.template_name, context)

            if not items:
                messages.error(request, "No se encontraron medicamentos con el formato esperado.")
                context["errores"] = errores
                return render(request, self.template_name, context)

            creados = 0
            actualizados = 0

            for item in items:
                defaults = {
                    "nombre": item.nombre,
                    "departamento": item.departamento,
                    "categoria": item.categoria,
                    "existencia": item.existencia,
                    "precio": item.precio,
                }
                if item.imagen_data:
                    ext = (item.imagen_ext or "png").lower().replace(".", "")
                    filename = f"{item.clave}.{ext}"
                    defaults["imagen"] = ContentFile(item.imagen_data, name=filename)

                obj, created = MedicamentoCatalogo.objects.update_or_create(
                    clave=item.clave, defaults=defaults
                )
                creados += int(created)
                actualizados += int(not created)

            messages.success(
                request,
                f"Importación completada: {creados} creados, {actualizados} actualizados.",
            )
            context.update(
                {
                    "creados": creados,
                    "actualizados": actualizados,
                    "total": len(items),
                    "errores": errores,
                }
            )
        return render(request, self.template_name, context)
